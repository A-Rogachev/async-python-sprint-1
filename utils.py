import os
import shutil
from typing import Any

import openpyxl
import requests
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def internet_connection_is_available() -> bool:
    """
    Проверка подключения к интернету.
    """
    try:
        requests.get('https://google.com')
        return True
    except Exception:
        return False


CITIES = {
    'MOSCOW': (
        'https://code.s3.yandex.net/async-module/moscow-response.json'
    ),
    'PARIS': (
        'https://code.s3.yandex.net/async-module/paris-response.json'
    ),
    'LONDON': (
        'https://code.s3.yandex.net/async-module/london-response.json'
    ),
    'BERLIN': (
        'https://code.s3.yandex.net/async-module/berlin-response.json'
    ),
    'BEIJING': (
        'https://code.s3.yandex.net/async-module/beijing-response.json'
    ),
    'KAZAN': (
        'https://code.s3.yandex.net/async-module/kazan-response.json'
    ),
    'SPETERSBURG': (
        'https://code.s3.yandex.net/async-module/spetersburg-response.json'
    ),
    'VOLGOGRAD': (
        'https://code.s3.yandex.net/async-module/volgograd-response.json'
    ),
    'NOVOSIBIRSK': (
        'https://code.s3.yandex.net/async-module/novosibirsk-response.json'
    ),
    'KALININGRAD': (
        'https://code.s3.yandex.net/async-module/kaliningrad-response.json'
    ),
    'ABUDHABI': (
        'https://code.s3.yandex.net/async-module/abudhabi-response.json'
    ),
    'WARSZAWA': (
        'https://code.s3.yandex.net/async-module/warszawa-response.json'
    ),
    'BUCHAREST': (
        'https://code.s3.yandex.net/async-module/bucharest-response.json'
    ),
    'ROMA': (
        'https://code.s3.yandex.net/async-module/roma-response.json'
    ),
    'CAIRO': (
        'https://code.s3.yandex.net/async-module/cairo-response.json'
    ),
    'GIZA': (
        'https://code.s3.yandex.net/async-module/giza-response.json'
    ),
    'MADRID': (
        'https://code.s3.yandex.net/async-module/madrid-response.json'
    ),
    'TORONTO': (
        'https://code.s3.yandex.net/async-module/toronto-response.json'
    ),
}

CITIES_NAMES_TRANSLATION: dict[str, str] = {
    'MOSCOW': 'Москва',
    'PARIS': 'Париж',
    'LONDON': 'Лондон',
    'BERLIN': 'Берлин',
    'BEIJING': 'Пекин',
    'KAZAN': 'Казань',
    'SPETERSBURG': 'Санкт-Петербург',
    'VOLGOGRAD': 'Волгоград',
    'ROMA': 'Рим',
    'NOVOSIBIRSK': 'Новосибирск',
    'KALININGRAD': 'Калининград',
    'ABUDHABI': 'Абу-Даби',
    'WARSZAWA': 'Варшава',
    'BUCHAREST': 'Бухарест',
    'CAIRO': 'Каир',
}


excel_report_table_settings: dict[str, Any] = {
    'bold_font': Font(bold=True),
    'thin_border': Border(
        left=Side(style='thin', color='325180'),
        right=Side(style='thin', color='325180'),
        top=Side(style='thin', color='325180'),
        bottom=Side(style='thin', color='325180'),
    ),
    'center_alignment': Alignment(horizontal='center'),
    'color_fill': PatternFill(
        start_color='C7E4E2',
        end_color='C1E4E7',
        fill_type='solid',
    ),
    'sheet_title': 'Анализ погоды',
    'sheet_names': {
        'A1': 'Город/день',
        'B1': '',
        'C1': '26-05',
        'D1': '27-05',
        'E1': '28-05',
        'F1': '29-05',
        'G1': '30-05',
        'H1': 'Среднее',
        'I1': 'Рейтинг',
    },
    'first_column_width': 20,
    'second_column_width': 25,
}

MIN_MAJOR_PYTHON_VER = 3
MIN_MINOR_PYTHON_VER = 9


class ReportExcelTable:
    """
    Класс отчета о погодных условиях в формате Excel (.xlsx).
    """

    def __init__(
        self,
        file_path: str,
        settings: dict[str, Any],
        records_amount: int,
    ) -> None:
        """
        Получение настроек для отчета в формате Excel.
        """
        self.records_amount: int = records_amount
        self.file_path: str = file_path
        self.thin_border: Border | None = settings.get('thin_border')
        self.bold_font: Font | None = settings.get('bold_font')
        self.center_alignment: Alignment | None = settings.get(
            'center_alignment'
        )
        self.title: Any = settings.get('sheet_title')
        self.sheet_names: dict[str, Any] | Any = settings.get('sheet_names')
        self.color_fill: PatternFill | None = settings.get('color_fill')
        self.first_column_width: int | None = settings.get(
            'first_column_width'
        )
        self.second_column_width: int | None = settings.get(
            'second_column_width'
        )

    def create_and_setup_new_excel_file(self) -> None:
        """
        Создает новый файл и делает базовую настройку.
        """
        wb: openpyxl.Workbook = openpyxl.Workbook()
        wb.active.title = self.title
        sheet = wb[self.title]
        for key, value in self.sheet_names.items():
            sheet[key] = value
            sheet[key].font = self.bold_font
            sheet[key].alignment = self.center_alignment
        sheet.column_dimensions['A'].width = self.first_column_width
        sheet.column_dimensions['B'].width = self.second_column_width

        for i in range(3, (self.records_amount + 1) * 2, 2):
            for cell in (key[0] for key in self.sheet_names.keys()):
                sheet[f'{cell}{i}'].fill = self.color_fill
        for k in range(1, (self.records_amount + 1) * 2):
            for cell in (key[0] for key in self.sheet_names.keys()):
                sheet[f'{cell}{k}'].border = self.thin_border

        wb.save(self.file_path)
        wb.close()


def check_python_version():
    import sys

    if (
        sys.version_info.major < MIN_MAJOR_PYTHON_VER
        or sys.version_info.minor < MIN_MINOR_PYTHON_VER
    ):
        raise Exception(
            'Please use python version >= {}.{}'.format(
                MIN_MAJOR_PYTHON_VER, MIN_MINOR_PYTHON_VER
            )
        )


def get_url_by_city_name(city_name):
    try:
        return CITIES[city_name]
    except KeyError:
        raise Exception('Please check that city {} exists'.format(city_name))


def create_new_folders(folder_names: tuple[str, ...]) -> None:
    """
    Создает новые директории; в случае если директория уже существует,
    рекурсивно удаляет вложенные папки и файлы, также саму директорию.
    """
    for folder_name in folder_names:
        try:
            os.mkdir(str(folder_name))
        except FileExistsError:
            shutil.rmtree(folder_name)
            os.mkdir(folder_name)
